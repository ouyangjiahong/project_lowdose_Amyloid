import argparse
from scipy import io as sio
import numpy as np
import os
import dicom
import nibabel as nib
import openpyxl as px
import sys
import pdb
from data_preparation_classification_tools import *

parser = argparse.ArgumentParser(description='')
parser.add_argument('--set', dest='set', default='train', help='train, test, test_subj')
parser.add_argument('--subj_id', dest='subj_id', type=int, default='1355')
parser.add_argument('--dir_label_ori', dest='dir_label_ori', default='diagnosis_status_all.xlsx', help='path of the label xlsx')
parser.add_argument('--dir_label_dst', dest='dir_label_dst', default='classification_label.npy', help='path of the data label')
parser.add_argument('--dir_data_ori', dest='dir_data_ori', default='../data/', help='folder of original data')
parser.add_argument('--dir_data_dst', dest='dir_data_dst', default='data/classification/', help='folder of jpg classification data')
parser.add_argument('--is_norm', dest='is_norm', action='store_true', help='divided by volume mean value)')
parser.set_defaults(is_norm=False)

args = parser.parse_args()
set = args.set
dir_label_ori = args.dir_label_ori
dir_label_dst = set + '_' + args.dir_label_dst
dir_data_ori = args.dir_data_ori
dir_data_raw = dir_data_ori + 'Amyloid/'
dir_data_res = dir_data_ori + 'Amyloid_Kevin_result/'
dir_data_dst = args.dir_data_dst
norm = args.is_norm
subj_id = args.subj_id

dir_data_dst = dir_data_dst + set + '/'
if not os.path.exists(dir_data_dst):
    os.makedirs(dir_data_dst)
if set == 'test_subj':
    dir_data_dst = dir_data_dst + str(subj_id) + '/'
    if not os.path.exists(dir_data_dst):
        os.makedirs(dir_data_dst)

'''
labels
subj_number, idx, age, gender, diagnosis, consensus, fulldose 1, fulldose 2,
pet+mr 1, pet+mr 2, petonly 1, petonly 2
'''
label_xls = px.load_workbook(dir_label_ori)
sheet = label_xls.get_sheet_by_name(name='Sheet1')
label_dict = {}
num = 1
for row in sheet.iter_rows():
    if num == 1 or num == 2:    # title
        num += 1
        continue
    # pdb.set_trace()
    subj = str(sheet.cell(column=1, row=num).value)

    # full_dose label
    label1 = str(sheet.cell(column=7, row=num).value)
    label2 = str(sheet.cell(column=8, row=num).value)
    if label1 == 'P' and label2 == 'P':
        fulldose_label = 1
    elif label1 == 'P' or label2 == 'P':
        fulldose_label = 0.5
    else:
        fulldose_label = 0

    # pet+mr label
    label1 = str(sheet.cell(column=9, row=num).value)
    label2 = str(sheet.cell(column=10, row=num).value)
    if label1 == 'P' and label2 == 'P':
        pet_mr_label = 1
    elif label1 == 'P' or label2 == 'P':
        pet_mr_label = 0.5
    else:
        pet_mr_label = 0

    # petonly label
    label1 = str(sheet.cell(column=11, row=num).value)
    label2 = str(sheet.cell(column=12, row=num).value)
    if label1 == 'P' and label2 == 'P':
        petonly_label = 1
    elif label1 == 'P' or label2 == 'P':
        petonly_label = 0.5
    else:
        petonly_label = 0

    label_dict[subj] = [fulldose_label, pet_mr_label, petonly_label]
    num += 1
print(label_dict)

'''
dataset
'''
filename_init = ''
# all_set = [1350, 1355, 1375, 1726, 1732, 1750, 1758, 1762, 1785, 1789,
#             1791, 1816, 1827, 1838, 1905, 1907, 1923, 1947, 1961, 1965, 1978,
#             2014, 2016, 2063, 2152, 2157, 2185, 2214, 2304, 2314, 2317,
#             2376, 2414, 2416, 2425, 2427, 2482, 2511, 2516, 50767]

# train_set = [1350, 1726, 1750, 1758, 1762, 1785, 1791, 1827, 1838, 1905, 1907,
#             1978, 2014, 2016, 2157, 2214, 2304, 2317, 2376, 2427, 2414, 1961,
#             2185, 2152, 1375, 1789, 1816, 1965, 2314, 2511, 2416, 2425]
# test_set = [1355, 1732, 1947, 2482, 2516, 2063, 1923, 50767]

train_set = [1350, 1726, 1750, 1762, 1785, 1791, 1827, 1838, 1905, 1907,
            1978, 2014, 2016, 2157, 2214, 2304, 2317, 2376, 2427, 2414, 1961,
            2185, 2152, 1789, 1816, 1965, 2314, 2511, 2416, 2482]
test_set = [1355, 1732, 1947, 2516, 2063, 50767, 1375, 1758, 1923, 2425]

if set == 'test':
    using_set = test_set
elif set == 'train':
    using_set = train_set
else:
    using_set = [args.subj_id]

list_subject = [str(x) for x in using_set if os.path.isdir(os.path.join(dir_data_ori, 'Amyloid', str(x)))]
print('generating subject list:')
print(list_subject)

list_dataset_train = []
list_dataset_label = []
filename_fulldose = 'pet_nifti/500_.nii.gz'
filename_pet_mr = 'mc.nii.gz'
filename_petonly = 'petonly.nii.gz'

for subject_id in list_subject:
    dir_subject = os.path.join(dir_data_raw, subject_id)
    list_dataset_train.append({'gt':os.path.join(dir_subject, filename_fulldose)})
    list_dataset_label.append(label_dict[subject_id][0])
    # dir_subject = os.path.join(dir_data_res, subject_id)
    # list_dataset_train.append({'gt':os.path.join(dir_subject, filename_pet_mr)})
    # list_dataset_label.append(label_dict[subject_id][1])
    # list_dataset_train.append({'gt':os.path.join(dir_subject, filename_petonly)})
    # list_dataset_label.append(label_dict[subject_id][2])

num_dataset_train = len(list_dataset_train)
print('process {0} data description'.format(num_dataset_train))
print(list_dataset_label)
# pdb.set_trace()

'''
augmentation
'''
list_augments = []
if set == 'train':
    num_augment_flipxy = 2
    num_augment_flipx = 2
    num_augment_flipy = 2
    num_augment_shiftx = 1
    num_augment_shifty = 1
else:
    num_augment_flipxy = 1
    num_augment_flipx = 1
    num_augment_flipy = 1
    num_augment_shiftx = 1
    num_augment_shifty = 1

for flipxy in range(num_augment_flipxy):
    for flipx in range(num_augment_flipx):
        for flipy in range(num_augment_flipy):
            for shiftx in range(num_augment_shiftx):
                for shifty in range(num_augment_shifty):
                    augment={'flipxy':flipxy,'flipx':flipx,'flipy':flipy,'shiftx':shiftx,'shifty':shifty}
                    list_augments.append(augment)
num_augment=len(list_augments)
print('will augment data with {0} augmentations'.format(num_augment))

'''
file loading related
'''
ext_data = 'npz'

'''
generate train data
'''
label_list = []
index_sample_total = 0
# pdb.set_trace()
for index_data in range(num_dataset_train):
    # directory
    # load data ground truth
    path_train_gt = list_dataset_train[index_data]['gt']
    label = list_dataset_label[index_data]
    data_train_gt = prepare_data_from_nifti(path_train_gt, list_augments, scale_by_norm=norm)

    # export
    index_sample_total, label_list = export_data_to_jpg(data_train_gt, dir_data_dst,
                                        label, label_list, index_sample_total, ext_data)

with open(dir_label_dst,'w') as file_input:
    np.save(file_input, label_list)
